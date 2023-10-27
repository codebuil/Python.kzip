
import calendar
import openpyxl
import sys

def create_year_calendar(year):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active

        days_of_week = list(calendar.day_abbr)  # Abreviações dos dias da semana
        months = list(calendar.month_abbr)[1:]  # Abreviações dos meses

        # Escreve os cabeçalhos dos dias da semana
        for i, day in enumerate(days_of_week):
            sheet.cell(row=1, column=i + 2, value=day)

        for i, month in enumerate(months):
            sheet.cell(row=2 + i*8, column=1, value=month)  # Escreve os meses

            # Obtém o número de dias em cada mês do ano
            month_days = calendar.monthrange(year, i+1)[1]

            # Cria o calendário para o mês atual
            cal = calendar.monthcalendar(year, i+1)

            for week_num, week in enumerate(cal):
                for day_num, day in enumerate(week):
                    if day != 0:
                        sheet.cell(row=week_num + 2 + i*8, column=day_num + 2, value=day)

        output_file = f"calendario_{year}.xlsx"
        wb.save(output_file)
        print(f"Calendário do ano {year} gerado com sucesso no arquivo: {output_file}")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

if __name__ == "__main__":
    print("\x1bc\x1b[43;30m")

    if len(sys.argv) != 2:
        print("Uso: python create_year_calendar.py <ano>")
    else:
        year = int(sys.argv[1])
        create_year_calendar(year)
